from openpyxl import Workbook
from openpyxl.styles import Font,Border,Side,PatternFill

wb = Workbook()

ws_users = wb.active
ws_users.title = 'Users'

wb.create_sheet("Statistics")
ws_statistics = wb["Statistics"]

ws_users.append(["Name","Age","City"])

users = [
    ["Serhii", 24, "Torun"],
    ["Alex", 30, "Bydgoszcz"],
    ["John", 18, "Warszawa"],
    ["Mike", 27, "Gdansk"],
]

for row_num,user in enumerate(users,start=2):
    for col,u in enumerate(user,start=1):
        ws_users.cell(row=row_num,column=col).value = u
ws_statistics.append(["Metric","Value"])
ws_statistics.column_dimensions["A"].width = 20
stats_name = ["Total users","Average age","Maximum age","Minimum age"]
for col,stat in enumerate(stats_name,start=2):
    ws_statistics.cell(row=col,column=1).value = stat
numbers_colums = []
users_len = len(users)
numbers_colums.append(users_len)

total_age = sum(user[1] for user in users)
average_age = total_age/users_len
numbers_colums.append(average_age)

max_age = max(user[1] for user in users)
numbers_colums.append(max_age)

min_age = min(user[1] for user in users)
numbers_colums.append(min_age)


for row_num,numbers_col in enumerate(numbers_colums,start=2):
    ws_statistics.cell(row=row_num,column=2).value = numbers_col


for user_title in ws_users[1]:
    user_title.font = Font(bold=True)

for stats_title in ws_statistics[1]:
    stats_title.font = Font(bold=True)

thin = Side(style='thin')
border = Border(
    right=thin,
    left=thin,
    top=thin,
    bottom=thin,
)
for row in ws_users.iter_rows():
    for cell in row:
        cell.border = border

for row in ws_statistics.iter_rows():
    for cell in row:
        cell.border = border

fill = PatternFill(fill_type='solid',fgColor='D9EAF7')
for row in ws_users[1]:
    row.fill = fill

for row in ws_statistics[1]:
    row.fill = fill

ws_users.column_dimensions["A"].width = 15
ws_users.column_dimensions["B"].width = 10
ws_users.column_dimensions["C"].width = 20

ws_statistics.column_dimensions["A"].width = 20
ws_statistics.column_dimensions["B"].width = 15

ws_users.freeze_panes = "A2"
ws_statistics.freeze_panes = "A2"

ws_users.auto_filter.ref = ws_users.calculate_dimension()


ws_statistics["B3"].number_format = "0.00"


wb.save("excel_report_final.xlsx")