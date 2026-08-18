from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
wb = load_workbook("ExelFolder/users.xlsx")
ws = wb.active

fill = PatternFill(fill_type="solid",  start_color="FFFC0000", end_color="FFFC0000")


rule = CellIsRule(
    operator="greaterThan",
    formula=[25],
    fill=fill,
)

ws.conditional_formatting.add("B2:B5",rule)
wb.save("ExelFolder/users_rule.xlsx")

