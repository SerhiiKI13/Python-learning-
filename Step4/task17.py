from openpyxl import load_workbook
wb = load_workbook('users3.xlsx')
ws = wb.active
print(ws['A1'].value)
print(ws['B2'].value)
print(ws['C4'].value)
