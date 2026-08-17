from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "Name"
ws["B1"] = "Age"
ws["C1"] = "City"

wb.save("users.xlsx")

print("Excel created")