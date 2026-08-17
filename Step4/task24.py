from openpyxl import load_workbook
wb = load_workbook('users.xlsx')
ws = wb.active

empty_rows = []
for row in ws.iter_rows(min_row=2):
    if row[0].value is None and row[1].value is None and row[2].value is None:
        empty_rows.append(row[0].row)
for row_num in reversed(empty_rows):
        ws.delete_rows(row_num)

wb.save('users.xlsx')