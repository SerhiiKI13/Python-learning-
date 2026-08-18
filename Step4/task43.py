from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
wb = Workbook()
ws = wb.active
users = [
    ("Serhii1", 18),
    ("Serhii2", 17),
    ("Serhii3", 30),
    ("Serhii4", 29),
    ("Serhii5", 24),
    ("Anton", 5),
]

ws.append(['Name',"Age"])

for user in users:
    ws.append(user)

fill = PatternFill(fill_type='solid',start_color="FFFC0000", end_color="FFFC0000")

rule = CellIsRule(
    operator="greaterThan",
    formula=["25"],
    fill=fill,
)
fill2 = PatternFill(fill_type='solid',start_color="FFFC9612", end_color="FFFC9612")
rule2 = CellIsRule(
    operator="lessThan",
    formula=["20"],
    fill=fill2,
)
ws.conditional_formatting.add(f"B2:B{ws.max_row}",rule)
ws.conditional_formatting.add(f"B2:B{ws.max_row}",rule2)
wb.save('users_new_task43.xlsx')