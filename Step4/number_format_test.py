from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.append(["Value","Percent","Price"])
nums = [1000,0.25,1235.5]
for col_num,num in enumerate(nums,start=1):
    ws.cell(row = 2, column = col_num).value = num

formats = ["0", "0%", "#,##0.00"]
for row_col,format in enumerate(formats,start=1):
    ws.cell(row = 2, column = row_col).number_format = format


wb.save("number_format_test.xlsx")