from openpyxl import load_workbook
wb = load_workbook('users2.xlsx')
ws = wb.active
ws['C2'] = 'Bydgoszcz'
wb.save('users2.xlsx')