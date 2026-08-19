from openpyxl import load_workbook
wb = load_workbook("formula_report.xlsx")
ws = wb.active
ws_stats = wb["Statistics"]

ws_stats['B2'] = f"=COUNTA(Users!A2:A{ws.max_row})"
ws_stats['B3'] = f"=AVERAGE(Users!B2:B{ws.max_row})"
ws_stats['B4'] = f"=MAX(Users!B2:B{ws.max_row})"
ws_stats['B5'] = f"=MIN(Users!B2:B{ws.max_row})"




wb.save("formula_report.xlsx")