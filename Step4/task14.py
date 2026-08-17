from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['C1'] = 'City'
ws['A2'] = 'Serhii'
ws['B2'] = 24
ws['C2'] = 'Torun'
ws['A3'] = 'Alex'
ws['B3'] = 25
ws['C3'] = 'Bydgoszcz'
ws['A4'] = 'John'
ws['B4'] = 30
ws['C4'] = 'Warszawa'
wb.save('users.xlsx')