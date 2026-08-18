from datetime import datetime

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.append([datetime.now(),"Event"])
ws.append([datetime.now(),"Python Course"])
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 20
ws["A1"].number_format = "DD.MM.YYYY"
ws["A2"].number_format = "DD.MM.YYYY HH:MM"
wb.save("datetime_test.xlsx")
