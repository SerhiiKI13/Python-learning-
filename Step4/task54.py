import csv
from openpyxl import Workbook
from openpyxl.styles import Font,Border,Side,PatternFill

fill = PatternFill(fill_type='solid',fgColor='D9EAF7')

wb = Workbook()
ws = wb.active
ws.title = "Users"
ws.append(["Name", "Age","City"])
with open("users_dirty.csv","r",encoding="utf-8") as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        try:
            age = int(row["Age"])
        except ValueError:
            continue
        if age >= 25:
            ws.append([row["Name"],age,row["City"]])

thin = Side(style= "thin",)
border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin,
)
for row in ws[1]:
    row.font = Font(bold=True)
    row.fill = fill


for row in ws.iter_rows():
    for cell in row:
        cell.border = border

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 20

ws.freeze_panes = "A2"

ws.auto_filter.ref = ws.calculate_dimension()

wb.create_sheet("Statistics")
ws_stats = wb["Statistics"]

ws_stats.append(["Metric","Value"])
statistics = ["Total users","Average age","Maximum age","Minimum age"]

for row,sts in enumerate(statistics,start=2):
    ws_stats.cell(row=row,column=1).value = sts

ws_stats["B2"] = (f"=COUNTA(Users!A2:A{ws.max_row})")
ws_stats["B3"] = (f"=AVERAGE(Users!B2:B{ws.max_row})")
ws_stats["B4"] = (f"=MAX(Users!B2:B{ws.max_row})")
ws_stats["B5"] = (f"=MIN(Users!B2:B{ws.max_row})")

for row in ws_stats[1]:
    row.font = Font(bold=True)
    row.fill = fill

for row in ws_stats.iter_rows():
    for cell in row:
        cell.border = border

ws_stats.column_dimensions["A"].width = 20
ws_stats.column_dimensions["B"].width = 15

ws_stats.freeze_panes = "A2"

ws_stats["B3"].number_format = "0.00"
wb.save("users_csv_report.xlsx")
