from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side,PatternFill
wb = load_workbook('dirty_users.xlsx')
ws = wb.active

thin = Side(style='thin')
border = Border(
    right=thin,
    top=thin,
    bottom=thin,
    left=thin,
)
fill = PatternFill(fill_type='solid',fgColor='D9EAF7')



empty_rows = []

for row in ws.iter_rows(min_row=2):
    if row[0].value is None:
        empty_rows.append(row[0].row)



for row in reversed(empty_rows):
  ws.delete_rows(row)

for row in ws.iter_rows(min_row=2):
    if row[2].value == "Torun":
        row[2].value = "Bydgoszcz"

for cell in ws[1]:
    cell.font = Font(bold=True)

for row in ws.iter_rows():
    for cell in row:
        cell.border = border

for cell in ws[1]:
    cell.fill = fill

ws.auto_filter.ref = "A1:C5"
ws.freeze_panes = "A2"

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 20

wb.save('clean_users.xlsx')


