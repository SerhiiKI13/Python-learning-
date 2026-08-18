from openpyxl import load_workbook
wb = load_workbook('users3.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if row[2].value == 'Torun':
        row[2].value = 'Bydgoszcz'
wb.save('users3.xlsx')