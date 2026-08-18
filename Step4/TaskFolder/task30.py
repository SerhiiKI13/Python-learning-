from openpyxl import load_workbook
from openpyxl.styles import Border,Side
wb = load_workbook('users.xlsx')
ws = wb.active
thin = Side(style='thin')
border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin,
)

for row in ws.iter_rows():
    for cell in row:
        cell.border = border

wb.save('users.xlsx')