from openpyxl import load_workbook
wb = load_workbook("cell_loop.xlsx")
ws = wb.active
print(ws.max_row)
print(ws.max_column)
print(ws.calculate_dimension())