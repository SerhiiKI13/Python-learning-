from openpyxl import load_workbook
wb = load_workbook("sheets_test2.xlsx")
ws = wb["Statistics"]
ws.title = "Stats"
ws_report = wb["Report"]
ws_report.title = "Final Report"
print(wb.sheetnames)

wb.save("renamed_sheets_test2.xlsx")