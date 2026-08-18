from openpyxl import load_workbook
wb = load_workbook("formula_test.xlsx")
ws = wb.active
print(ws['A4'].value)
wb1 = load_workbook("formula_test.xlsx", data_only=True)
ws1 = wb1.active
print(ws1['A4'].value)