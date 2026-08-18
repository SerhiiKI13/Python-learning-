from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "Name"
ws.cell(row=2, column=1).value = "Serhii"
ws.cell(row=3, column=1).value = "Alex"

ws.cell(row=1, column=2).value = "Age"
ws.cell(row=2, column=2).value = 25
ws.cell(row=3, column=2).value = 24

ws.cell(row=1, column=3).value = "City"
ws.cell(row=2, column=3).value = "Torun"
ws.cell(row=3, column=3).value = "Bydgoszcz"

wb.save("cell_test.xlsx")