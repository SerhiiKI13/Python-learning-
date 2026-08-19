from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Users"
ws2 = wb.create_sheet("Statistics")
ws3 = wb.create_sheet("Report")


wb.save("sheets_test.xlsx")