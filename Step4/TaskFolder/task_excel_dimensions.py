from openpyxl import load_workbook
wb = load_workbook('users3.xlsx')
ws = wb.active
print(ws.max_row)
print(ws.max_column)