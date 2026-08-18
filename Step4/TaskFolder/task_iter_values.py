from openpyxl import load_workbook
wb = load_workbook("cell_loop.xlsx")
ws = wb.active

for row in ws.iter_rows(
    min_row=3,
    max_row=4,
    min_col=1,
    max_col=2,
    values_only=True
):
    print(row)
