from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.merge_cells('A1:C1')
ws['A1'] = 'USER REPORT'
wb.save('merge_test.xlsx')