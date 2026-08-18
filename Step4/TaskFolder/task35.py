from openpyxl import load_workbook
wb = load_workbook('merge_test.xlsx')
ws = wb.active
ws.unmerge_cells('A1:C1')
ws['A1'] = 'name'
ws['B1'] = 'age'
ws['C1'] = 'city'

wb.save('unmerge_test.xlsx')