from openpyxl import load_workbook
wb = load_workbook('users3.xlsx')
ws = wb.active

for row in ws["A1:C3"]:
    values = []
    for cell in row:
        values.append(cell.value)
    values = [str(value) for value in values]
    print(" ".join(values))