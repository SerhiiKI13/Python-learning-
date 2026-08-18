from openpyxl import load_workbook
wb= load_workbook('users2.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if row[1].value > 24:
      print(row[0].value)
