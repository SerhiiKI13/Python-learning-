from openpyxl import load_workbook
from openpyxl.styles import PatternFill
wb = load_workbook('users.xlsx')
ws = wb.active

fill = PatternFill(fill_type='solid',fgColor='D9EAF7')

for cell in ws[1]:
    cell.fill = fill

wb.save('users.xlsx')