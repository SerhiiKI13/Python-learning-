from openpyxl import load_workbook
wb = load_workbook('users3.xlsx')
ws = wb.active

ws.insert_cols(2)
ws['B1'] = "Country"
for row in ws.iter_rows(min_row=2):
    row[1].value = "Poland"

wb.save("insert_cols_test.xlsx")