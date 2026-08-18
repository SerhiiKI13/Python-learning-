from openpyxl import load_workbook
from openpyxl.styles import Alignment
wb = load_workbook('users.xlsx')
ws = wb.active

for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center')

wb.save('users.xlsx')