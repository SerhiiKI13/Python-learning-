from openpyxl import load_workbook
wb = load_workbook('users3.xlsx')
ws = wb.active
ws.insert_rows(2)
ws["A2"] = "Mike"
ws["B2"] = 27
ws["C2"] = "Gdansk"

wb.save('insert_rows_test.xlsx')