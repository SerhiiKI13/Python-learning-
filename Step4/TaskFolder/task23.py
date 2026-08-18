from openpyxl import load_workbook
wb = load_workbook('users2.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if row[0].value == 'Alex':
       rowNumber = int(row[0].row)
       print(rowNumber)
       ws.delete_rows(rowNumber)
wb.save('users2.xlsx')
